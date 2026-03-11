# server/main.py
from fastapi import FastAPI, Depends, HTTPException, status
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session
from fastapi.security import OAuth2PasswordRequestForm
from auth import get_current_user, require_admin, require_superadmin, hash_password
from models import User, Test, Assignment, Result, Question, Option, ExitLog
from models import Response as DBResponse
from datetime import datetime
from datetime import date
from schemas import TestSubmission
from database import engine, Base, SessionLocal, get_db
from auth import verify_password, create_access_token
from schemas import Token
from schemas import UserCreate
from typing import Optional  # <-- ADDED for optional query parameters
from scoring.disc import score_disc
from scoring.speed import score_speed
from scoring.temperament import score_temperament
from sqlalchemy.orm import joinedload  # to load options efficiently
from schemas import UserUpdate
from scoring.memory import score_memory
import secrets
import string
from scoring.logic import score_logic
from sqlalchemy import func, desc
from scoring.leadership import score_leadership
import csv
import io
from typing import List
from fastapi import UploadFile, File, Form
import openpyxl
import xlrd
from openpyxl.utils import get_column_letter
import csv
from fastapi.responses import StreamingResponse
import io
from weasyprint import HTML
from fastapi.responses import Response
from datetime import datetime
from fastapi import HTTPException, Response
from fastapi.responses import StreamingResponse
from weasyprint import HTML, CSS
from sqlalchemy.orm import Session
from typing import Dict, Any


# Create tables
Base.metadata.create_all(bind=engine)

app = FastAPI()

# CORS settings (unchanged)
origins = [
    "http://localhost:5173",
    "http://127.0.0.1:5173"
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Helper for max score per test (optional, used in charts)
def get_max_score(test_code):
    return {
        "DISC": 24,
        "SPEED": 100,
        "TEMP": 168,
        "MEM": 100,
        "LOGIC": 100,
    }.get(test_code, 100)

@app.get("/")
def read_root():
    return {"message": "Hello from Python Backend!"}

# --- LOGIN ROUTE (unchanged) ---
@app.post("/login", response_model=Token)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    user = db.query(User).filter(User.username == form_data.username).first()
    if not user:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect username or password",
        )
    if not verify_password(form_data.password, user.password_hash):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect username or password",
        )
    access_token = create_access_token(data={"sub": user.username})
    return {"access_token": access_token, "token_type": "bearer"}

@app.get("/users/me")
def read_users_me(current_user: User = Depends(get_current_user)):
    return {
        "username": current_user.username,
        "role": current_user.role,
        "id": current_user.id
    }

@app.post("/users/", status_code=201)
def create_user(
    user: UserCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)  # both superadmin and admin allowed
):
    # If current user is not superadmin, they cannot create admin/superadmin
    if current_user.role != "superadmin" and user.role in ["admin", "superadmin"]:
        raise HTTPException(
            status_code=403,
            detail="Only superadmin can create admin users"
        )
    
    db_user = db.query(User).filter(User.username == user.username).first()
    if db_user:
        raise HTTPException(status_code=400, detail="Username already registered")
    
    new_user = User(
        username=user.username,
        password_hash=hash_password(user.password),
        role=user.role,
        full_name=user.full_name,
        gender=user.gender,
        age=user.age,
        education=user.education,
        department=user.department,
        position=user.position
    )
    db.add(new_user)
    db.commit()
    db.refresh(new_user)
    return {"message": "User created successfully", "username": new_user.username}

@app.get("/users/")
def get_users(db: Session = Depends(get_db), admin: User = Depends(require_admin)):
    users = db.query(User).all()
    return [
        {
            "id": u.id, 
            "username": u.username, 
            "role": u.role,
            "full_name": u.full_name,
            "department": u.department,
            "position": u.position
        } 
        for u in users
    ]
    
@app.post("/admin/reset-password/{user_id}")
def reset_password(
    user_id: int,
    db: Session = Depends(get_db),
    superadmin: User = Depends(require_superadmin)
):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    # Generate a random 10-character password
    alphabet = string.ascii_letters + string.digits
    new_password = ''.join(secrets.choice(alphabet) for _ in range(10))

    # Hash and update
    user.password_hash = hash_password(new_password)
    db.commit()

    return {"new_password": new_password}

# ---------- ADDED: Get single user by ID ----------
@app.get("/users/{user_id}")
def get_user(user_id: int, db: Session = Depends(get_db), superadmin: User = Depends(require_superadmin)):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    return {
        "id": user.id,
        "username": user.username,
        "role": user.role,
        "full_name": user.full_name,
        "gender": user.gender,
        "age": user.age,
        "education": user.education,
        "department": user.department,
        "position": user.position
    }

@app.put("/users/{user_id}")
def update_user(
    user_id: int,
    user_update: UserUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    # If current user is not superadmin, they cannot change roles
    if current_user.role != "superadmin" and user_update.role is not None:
        raise HTTPException(
            status_code=403,
            detail="Only superadmin can change user roles"
        )

    # Also prevent a non-superadmin from elevating someone to admin/superadmin
    if current_user.role != "superadmin" and user_update.role in ["admin", "superadmin"]:
        raise HTTPException(
            status_code=403,
            detail="Only superadmin can assign admin or superadmin roles"
        )

    # Update fields (same as before)
    if user_update.username is not None:
        if user_update.username != user.username:
            existing = db.query(User).filter(User.username == user_update.username).first()
            if existing:
                raise HTTPException(status_code=400, detail="Username already taken")
        user.username = user_update.username
    if user_update.full_name is not None:
        user.full_name = user_update.full_name
    if user_update.gender is not None:
        user.gender = user_update.gender
    if user_update.age is not None:
        user.age = user_update.age
    if user_update.education is not None:
        user.education = user_update.education
    if user_update.department is not None:
        user.department = user_update.department
    if user_update.position is not None:
        user.position = user_update.position
    if user_update.role is not None:
        user.role = user_update.role
    if user_update.password is not None and user_update.password != "":
        user.password_hash = hash_password(user_update.password)

    db.commit()
    db.refresh(user)
    return {"message": "User updated successfully", "user": user}

@app.delete("/users/{user_id}")
def delete_user(
    user_id: int,
    db: Session = Depends(get_db),
    superadmin: User = Depends(require_superadmin)
):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    # Delete related records (cascade manually)
    db.query(ExitLog).filter(ExitLog.user_id == user_id).delete()
    db.query(Response).filter(Response.user_id == user_id).delete()
    db.query(Result).filter(Result.user_id == user_id).delete()
    db.query(Assignment).filter(Assignment.user_id == user_id).delete()
    db.delete(user)
    db.commit()

    return {"message": "User deleted successfully"}

# ---------- MODIFIED: Assignments endpoint with optional user_id filter ----------
@app.get("/assignments/")
def get_assignments(
    user_id: Optional[int] = None,  # <-- ADDED query parameter
    db: Session = Depends(get_db), 
    admin: User = Depends(require_admin)
):
    query = db.query(Assignment)
    if user_id is not None:
        query = query.filter(Assignment.user_id == user_id)
    assignments = query.all()
    result = []
    for a in assignments:
        result.append({
            "id": a.id,
            "user_id": a.user_id,
            "username": a.user.username,
            "full_name": a.user.full_name,
            "test_id": a.test_id,
            "test_name": a.test.name,
            "status": a.status,
            "assigned_at": a.assigned_at
        })
    return result

@app.post("/assignments/", status_code=201)
def create_assignment(user_id: int, test_id: int, db: Session = Depends(get_db), admin: User = Depends(require_admin)):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    test = db.query(Test).filter(Test.id == test_id).first()
    if not test:
        raise HTTPException(status_code=404, detail="Test not found")
    existing = db.query(Assignment).filter(Assignment.user_id == user_id, Assignment.test_id == test_id).first()
    if existing:
        raise HTTPException(status_code=400, detail="Test already assigned to this user")
    new_assignment = Assignment(user_id=user_id, test_id=test_id)
    db.add(new_assignment)
    db.commit()
    return {"message": "Test assigned successfully"}

@app.get("/users/me/assignments")
def get_my_assignments(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    assignments = db.query(Assignment).filter(Assignment.user_id == current_user.id).all()
    result = []
    for a in assignments:
        result.append({
            "id": a.id,
            "test_id": a.test_id,
            "test_name": a.test.name,
            "test_code": a.test.code,
            "status": a.status,
            "pretest_completed": a.pretest_completed,
            "assigned_at": a.assigned_at
        })
    return result

@app.get("/assignments/{assignment_id}/start")
def start_test(assignment_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    assignment = db.query(Assignment).filter(Assignment.id == assignment_id).first()
    if not assignment:
        raise HTTPException(status_code=404, detail="Assignment not found")
    if assignment.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized")
    if assignment.status == "locked":
        raise HTTPException(status_code=403, detail="This test is locked due to a security violation. Please contact the administrator.")
    if assignment.status == "pending":
        assignment.status = "in_progress"
        db.commit()
    questions = db.query(Question).filter(Question.test_id == assignment.test_id).order_by(Question.order_index).all()
    output = []
    for q in questions:
        options_data = []
        for opt in q.options:
            options_data.append({
                "id": opt.id,
                "label": opt.label,
                "content": opt.content
            })
        output.append({
            "id": q.id,
            "content": q.content,
            "order": q.order_index,
            "options": options_data
        })
    return {
        "test_name": assignment.test.name,
        "time_limit": assignment.test.time_limit,
        "settings": assignment.test.settings,
        "questions": output
    }

@app.post("/assignments/{assignment_id}/submit")
def submit_test(
    assignment_id: int,
    submission: TestSubmission,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    # 1. Find the assignment
    assignment = db.query(Assignment).filter(Assignment.id == assignment_id).first()
    if not assignment or assignment.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized")
    if assignment.status != "in_progress":
        raise HTTPException(status_code=400, detail="Test is not in progress")

    # 2. Save all responses
    for ans in submission.answers:
        resp = DBResponse(
            user_id=current_user.id,
            test_id=assignment.test_id,
            assignment_id=assignment.id,
            question_id=ans["question_id"],
            selected_option_id=ans.get("option_id"),
            selection_type=ans.get("type", "single"),
        )
        db.add(resp)

    # 3. Score based on test code
    test_code = assignment.test.code
    score = 0
    details = None

    if test_code == "DISC":
        questions = (
            db.query(Question)
            .options(joinedload(Question.options))
            .filter(Question.test_id == assignment.test_id)
            .all()
        )
        details = score_disc(submission.answers, questions)
        score = len({ans["question_id"] for ans in submission.answers})  # 24

    elif test_code == "SPEED":
        questions = (
            db.query(Question)
            .options(joinedload(Question.options))
            .filter(Question.test_id == assignment.test_id)
            .all()
        )
        details = score_speed(submission.answers, questions)
        score = details["score"]

    elif test_code == "TEMP":
        questions = (
            db.query(Question)
            .options(joinedload(Question.options))
            .filter(Question.test_id == assignment.test_id)
            .all()
        )
        details = score_temperament(submission.answers, questions)
        # Optional: you could set score to something meaningful, e.g., sum of raw scores
        score = sum(details["raw_scores"].values()) if details["raw_scores"] else 0
    
    elif test_code == "MEM":
        questions = (
            db.query(Question)
            .options(joinedload(Question.options))
            .filter(Question.test_id == assignment.test_id)
            .all()
        )
        details = score_memory(submission.answers, questions)
        score = details["score"]
        
    elif test_code == "LOGIC":
        questions = (
            db.query(Question)
            .options(joinedload(Question.options))
            .filter(Question.test_id == assignment.test_id)
            .all()
        )
        details = score_logic(submission.answers, questions)
        score = details["score"]
        
    elif test_code == "LEAD":
        questions = (
        db.query(Question)
        .options(joinedload(Question.options))
        .filter(Question.test_id == assignment.test_id)
        .all()
    )
        
        details = score_leadership(submission.answers, questions)
        # For leadership, the score can be the sum of raw scores (optional)
        score = sum(details["raw_scores"].values())

    else:
        # Default: IQ, Memory, Logic, etc. (simple correct count)
        for ans in submission.answers:
            if ans.get("type", "single") == "single":
                option = db.query(Option).filter(Option.id == ans["option_id"]).first()
                if option and option.scoring_logic.get("correct"):
                    score += 1

    # 4. Create Result record
    result = Result(
        user_id=current_user.id,
        test_id=assignment.test_id,
        assignment_id=assignment.id,
        score=score,
        time_taken=submission.time_taken,
        details=details,
        completed_at=datetime.utcnow()
    )
    db.add(result)

    # 5. Mark assignment as completed
    assignment.status = "completed"
    db.commit()

    return {
        "message": "Test submitted successfully",
        "score": score,
        "test_type": test_code
    }

# ---------- MODIFIED: Results endpoint with optional user_id filter and user_id in response ----------
@app.get("/results/")
def get_results(
    user_id: Optional[int] = None,
    test_id: Optional[int] = None,           # new
    search: Optional[str] = None,            # new
    from_date: Optional[date] = None,        # new
    to_date: Optional[date] = None,          # new
    db: Session = Depends(get_db),
    superadmin: User = Depends(require_superadmin)
):
    query = db.query(Result).join(User).join(Test)  # need joins for filtering on related fields

    if user_id is not None:
        query = query.filter(Result.user_id == user_id)
    if test_id is not None:
        query = query.filter(Result.test_id == test_id)
    if from_date is not None:
        query = query.filter(Result.completed_at >= from_date)
    if to_date is not None:
        query = query.filter(Result.completed_at <= to_date)
    if search:
        # search in user.full_name and user.username
        query = query.filter(
            (User.full_name.ilike(f"%{search}%")) | (User.username.ilike(f"%{search}%"))
        )

    results = query.all()
    output = []
    for r in results:
        total_questions = db.query(Question).filter(Question.test_id == r.test_id).count()
        # Determine max possible score based on test code
        if r.test.code == "DISC":
            max_score = 24
        elif r.test.code == "SPEED":
            max_score = 100
        elif r.test.code == "MEM":
            max_score = 100
        elif r.test.code == "LOGIC":
            max_score = 100
        elif r.test.code == "LEAD":
            max_score = None   # No single max score for Leadership
        elif r.test.code == "TEMP":
            max_score = None   # No single max score for Temperament (we show primary trait)
        else:
            max_score = total_questions

        output.append({
            "id": r.id,
            "user_id": r.user_id,
            "username": r.user.username,
            "full_name": r.user.full_name,
            "test_name": r.test.name,
            "test_id": r.test_id,
            "score": r.score,
            "total_questions": total_questions,
            "max_score": max_score,          # <-- new field
            "time_taken": r.time_taken,
            "completed_at": r.completed_at,
            "details": r.details
        })
    return output

@app.post("/assignments/{assignment_id}/lock")
def lock_assignment(assignment_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    assignment = db.query(Assignment).filter(Assignment.id == assignment_id).first()
    if not assignment or assignment.user_id != current_user.id:
        raise HTTPException(status_code=404, detail="Assignment not found")
    assignment.status = "locked"
    db.commit()
    return {"message": "Assignment locked due to integrity violation"}

@app.post("/admin/assignments/{assignment_id}/unlock")
def unlock_assignment(assignment_id: int, db: Session = Depends(get_db), superadmin: User = Depends(require_superadmin)):
    assignment = db.query(Assignment).filter(Assignment.id == assignment_id).first()
    if not assignment:
        raise HTTPException(status_code=404, detail="Assignment not found")
    assignment.status = "in_progress"
    db.commit()
    return {"message": "Assignment unlocked"}

@app.get("/tests/")
def get_tests(db: Session = Depends(get_db), admin: User = Depends(require_admin)):
    return db.query(Test).all()

@app.post("/assignments/{assignment_id}/exit-log")
def log_exit(assignment_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    assignment = db.query(Assignment).filter(Assignment.id == assignment_id).first()
    if not assignment or assignment.user_id != current_user.id:
        raise HTTPException(status_code=404, detail="Assignment not found")
    log = ExitLog(user_id=current_user.id, assignment_id=assignment_id)
    db.add(log)
    db.commit()
    return {"message": "Exit logged"}

@app.get("/admin/locked-assignments")
def get_locked_assignments(db: Session = Depends(get_db), superadmin: User = Depends(require_superadmin)):
    assignments = db.query(Assignment).filter(Assignment.status == "locked").all()
    result = []
    for a in assignments:
        result.append({
            "id": a.id,
            "user_id": a.user_id,
            "username": a.user.username,
            "full_name": a.user.full_name,
            "test_name": a.test.name,
            "status": a.status,
            "assigned_at": a.assigned_at,
        })
    return result

@app.get("/admin/exit-logs")
def get_exit_logs(
    user_id: Optional[int] = None,
    assignment_id: Optional[int] = None,
    from_date: Optional[date] = None,
    to_date: Optional[date] = None,
    db: Session = Depends(get_db),
    superadmin: User = Depends(require_superadmin)
):
    query = db.query(ExitLog).join(User).join(Assignment)
    if user_id:
        query = query.filter(ExitLog.user_id == user_id)
    if assignment_id:
        query = query.filter(ExitLog.assignment_id == assignment_id)
    if from_date:
        query = query.filter(ExitLog.timestamp >= from_date)
    if to_date:
        query = query.filter(ExitLog.timestamp <= to_date)
    logs = query.order_by(ExitLog.timestamp.desc()).all()
    output = []
    for log in logs:
        output.append({
            "id": log.id,
            "user_id": log.user_id,
            "username": log.user.username,
            "full_name": log.user.full_name,
            "assignment_id": log.assignment_id,
            "test_name": log.assignment.test.name,
            "timestamp": log.timestamp
        })
    return output

@app.post("/admin/assignments/{assignment_id}/reset")
def reset_assignment(
    assignment_id: int,
    db: Session = Depends(get_db),
    admin: User = Depends(require_admin)          # both admin and superadmin can reset
):
    assignment = db.query(Assignment).filter(Assignment.id == assignment_id).first()
    if not assignment:
        raise HTTPException(status_code=404, detail="Assignment not found")

    # Delete all responses for this assignment
    db.query(Response).filter(Response.assignment_id == assignment_id).delete()
    # Delete the result for this assignment (if any)
    db.query(Result).filter(Result.assignment_id == assignment_id).delete()

    # Reset assignment status
    assignment.status = "pending"
    # Optionally reset the pretest flag so they see tutorial again
    assignment.pretest_completed = False

    db.commit()
    return {"message": "Assignment reset successfully"}

@app.post("/assignments/assign-all/{user_id}")
def assign_all_tests(
    user_id: int,
    db: Session = Depends(get_db),
    admin: User = Depends(require_admin)   # both admin and superadmin allowed
):
    # Check user exists
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    # Get all tests
    all_tests = db.query(Test).all()
    # Get already assigned test IDs for this user
    assigned_test_ids = {a.test_id for a in db.query(Assignment).filter(Assignment.user_id == user_id).all()}

    # Create assignments for missing tests
    created = []
    for test in all_tests:
        if test.id not in assigned_test_ids:
            new_assignment = Assignment(user_id=user_id, test_id=test.id)
            db.add(new_assignment)
            created.append(test.name)

    db.commit()
    return {
        "message": f"Assigned {len(created)} tests",
        "assigned_tests": created
    }
    
@app.get("/admin/stats/summary")
def get_stats_summary(db: Session = Depends(get_db), superadmin: User = Depends(require_superadmin)):
    total_participants = db.query(User).filter(User.role == "participant").count()
    total_assignments = db.query(Assignment).count()
    completed_assignments = db.query(Assignment).filter(Assignment.status == "completed").count()
    completion_rate = (completed_assignments / total_assignments * 100) if total_assignments else 0
    avg_score = db.query(func.avg(Result.score)).scalar() or 0

    return {
        "total_participants": total_participants,
        "total_assignments": total_assignments,
        "completed_assignments": completed_assignments,
        "completion_rate": round(completion_rate, 2),
        "average_score": round(avg_score, 2)
    }

@app.get("/admin/stats/tests")
def get_test_stats(db: Session = Depends(get_db), superadmin: User = Depends(require_superadmin)):
    tests = db.query(Test).all()
    result = []
    for test in tests:
        completed_count = db.query(Assignment).filter(Assignment.test_id == test.id, Assignment.status == "completed").count()
        avg_score = db.query(func.avg(Result.score)).filter(Result.test_id == test.id).scalar() or 0
        result.append({
            "test_id": test.id,
            "test_name": test.name,
            "completed_count": completed_count,
            "avg_score": round(avg_score, 2),
            "max_score": get_max_score(test.code)
        })
    return result

@app.get("/admin/stats/recent")
def get_recent_activity(limit: int = 10, db: Session = Depends(get_db), superadmin: User = Depends(require_superadmin)):
    recent_results = db.query(Result).order_by(desc(Result.completed_at)).limit(limit).all()
    output = []
    for r in recent_results:
        total_questions = db.query(Question).filter(Question.test_id == r.test_id).count()
        output.append({
            "id": r.id,
            "participant_name": r.user.full_name or r.user.username,
            "test_name": r.test.name,
            "score": r.score,
            "total_questions": total_questions,
            "completed_at": r.completed_at.isoformat() if r.completed_at else None
        })
    return output

@app.post("/assignments/{assignment_id}/complete-tutorial")
def complete_tutorial(
    assignment_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    assignment = db.query(Assignment).filter(Assignment.id == assignment_id).first()
    if not assignment or assignment.user_id != current_user.id:
        raise HTTPException(status_code=404, detail="Assignment not found")
    assignment.pretest_completed = True
    db.commit()
    return {"message": "Tutorial marked as completed"}

@app.post("/admin/users/bulk")
def bulk_create_users(
    file: UploadFile = File(...),
    assign_all: bool = Form(False),
    db: Session = Depends(get_db),
    admin: User = Depends(require_admin)
):
    filename = file.filename.lower()
    if not (filename.endswith('.csv') or filename.endswith('.xlsx') or filename.endswith('.xls')):
        raise HTTPException(status_code=400, detail="Only CSV or Excel files are allowed")

    contents = file.file.read()
    rows = []  # will be list of dicts

    if filename.endswith('.csv'):
        import csv
        import io
        decoded = contents.decode('utf-8-sig')
        csv_reader = csv.DictReader(io.StringIO(decoded))
        rows = list(csv_reader)

    elif filename.endswith('.xlsx'):
        from openpyxl import load_workbook
        from io import BytesIO
        wb = load_workbook(filename=BytesIO(contents), read_only=True)
        sheet = wb.active
        all_rows = list(sheet.iter_rows(values_only=True))
        if not all_rows:
            raise HTTPException(status_code=400, detail="Excel file is empty")
        headers = [str(cell) if cell is not None else '' for cell in all_rows[0]]
        rows = []
        for row in all_rows[1:]:
            row_dict = {}
            for idx, val in enumerate(row):
                if idx < len(headers):
                    row_dict[headers[idx]] = str(val) if val is not None else ''
            rows.append(row_dict)

    elif filename.endswith('.xls'):
        import xlrd
        from io import BytesIO
        book = xlrd.open_workbook(file_contents=contents)
        sheet = book.sheet_by_index(0)
        headers = [str(sheet.cell_value(0, col)) for col in range(sheet.ncols)]
        rows = []
        for row_idx in range(1, sheet.nrows):
            row_dict = {}
            for col_idx in range(sheet.ncols):
                val = sheet.cell_value(row_idx, col_idx)
                row_dict[headers[col_idx]] = str(val) if val else ''
            rows.append(row_dict)

    required_fields = ['username', 'password', 'full_name']
    results = {
        "total": len(rows),
        "success": 0,
        "failed": 0,
        "errors": []
    }
    created_users = []

    for row_num, row in enumerate(rows, start=2):
        missing = [f for f in required_fields if not row.get(f)]
        if missing:
            results["failed"] += 1
            results["errors"].append(f"Row {row_num}: Missing fields: {', '.join(missing)}")
            continue

        username = row['username'].strip()
        password = row['password'].strip()
        full_name = row['full_name'].strip()

        if db.query(User).filter(User.username == username).first():
            results["failed"] += 1
            results["errors"].append(f"Row {row_num}: Username '{username}' already exists")
            continue

        user_data = {
            'username': username,
            'full_name': full_name,
            'age': row.get('age') if row.get('age') else None,
            'gender': row.get('gender'),
            'education': row.get('education'),
            'department': row.get('department'),
            'position': row.get('position'),
            'role': 'participant'
        }
        if user_data['age']:
            try:
                user_data['age'] = int(user_data['age'])
            except ValueError:
                results["failed"] += 1
                results["errors"].append(f"Row {row_num}: Age must be a number")
                continue

        new_user = User(
            username=user_data['username'],
            password_hash=hash_password(password),
            role=user_data['role'],
            full_name=user_data['full_name'],
            age=user_data['age'],
            gender=user_data['gender'],
            education=user_data['education'],
            department=user_data['department'],
            position=user_data['position']
        )
        db.add(new_user)
        db.flush()
        created_users.append(new_user)
        results["success"] += 1

    if assign_all and created_users:
        all_tests = db.query(Test).all()
        for user in created_users:
            for test in all_tests:
                existing = db.query(Assignment).filter(
                    Assignment.user_id == user.id,
                    Assignment.test_id == test.id
                ).first()
                if not existing:
                    new_assignment = Assignment(user_id=user.id, test_id=test.id)
                    db.add(new_assignment)

    db.commit()
    return results

@app.get("/admin/export/results")
def export_results(
    test_id: Optional[int] = None,
    search: Optional[str] = None,
    from_date: Optional[date] = None,
    to_date: Optional[date] = None,
    db: Session = Depends(get_db),
    superadmin: User = Depends(require_superadmin)   # only superadmin can export
):
    # Build query (same as in get_results)
    query = db.query(Result).join(User).join(Test)
    if test_id:
        query = query.filter(Result.test_id == test_id)
    if from_date:
        query = query.filter(Result.completed_at >= from_date)
    if to_date:
        query = query.filter(Result.completed_at <= to_date)
    if search:
        query = query.filter(
            (User.full_name.ilike(f"%{search}%")) | (User.username.ilike(f"%{search}%"))
        )

    results = query.order_by(Result.completed_at.desc()).all()

    # Prepare CSV data
    output = io.StringIO()
    writer = csv.writer(output)
    # Write header
    writer.writerow([
        "Participant Username", "Participant Name", "Test", "Score", "Max Score",
        "Percentage", "Time Taken (s)", "Completed At"
    ])

    for r in results:
        # Determine max score based on test code
        if r.test.code == "DISC":
            max_score = 24
        elif r.test.code == "SPEED":
            max_score = 100
        elif r.test.code == "MEM":
            max_score = 100
        elif r.test.code == "LOGIC":
            max_score = 100
        elif r.test.code in ["TEMP", "LEAD"]:
            max_score = "N/A"
        else:
            max_score = db.query(Question).filter(Question.test_id == r.test_id).count()

        percentage = round((r.score / max_score * 100), 2) if isinstance(max_score, int) and max_score > 0 else "N/A"

        writer.writerow([
            r.user.username,
            r.user.full_name or "",
            r.test.name,
            r.score,
            max_score,
            percentage,
            r.time_taken,
            r.completed_at.strftime("%Y-%m-%d %H:%M") if r.completed_at else ""
        ])

    # Return as downloadable file
    response = StreamingResponse(iter([output.getvalue()]), media_type="text/csv")
    response.headers["Content-Disposition"] = "attachment; filename=results.csv"
    return response

@app.get("/admin/export/participant/{user_id}")
def export_participant_results(
    user_id: int,
    db: Session = Depends(get_db),
    superadmin: User = Depends(require_superadmin)
):
    # Get user
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    # Get all results for this user
    results = db.query(Result).filter(Result.user_id == user_id).order_by(Result.completed_at).all()

    # Prepare CSV
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Test", "Score", "Max Score", "Percentage", "Time Taken (s)", "Completed At"
    ])

    for r in results:
        # Determine max score
        if r.test.code == "DISC":
            max_score = 24
        elif r.test.code == "SPEED":
            max_score = 100
        elif r.test.code == "MEM":
            max_score = 100
        elif r.test.code == "LOGIC":
            max_score = 100
        elif r.test.code in ["TEMP", "LEAD"]:
            max_score = "N/A"
        else:
            max_score = db.query(Question).filter(Question.test_id == r.test_id).count()

        percentage = round((r.score / max_score * 100), 2) if isinstance(max_score, int) and max_score > 0 else "N/A"

        writer.writerow([
            r.test.name,
            r.score,
            max_score,
            percentage,
            r.time_taken,
            r.completed_at.strftime("%Y-%m-%d %H:%M") if r.completed_at else ""
        ])

    # Add participant info at top (optional) – but CSV doesn't have headers well for that.
    # We'll just return the table.

    response = StreamingResponse(iter([output.getvalue()]), media_type="text/csv")
    response.headers["Content-Disposition"] = f"attachment; filename={user.username}_results.csv"
    return response

@app.get("/admin/export/participant/{user_id}/pdf")
def export_participant_pdf(
    user_id: int,
    db: Session = Depends(get_db),
    superadmin: User = Depends(require_superadmin)
):
    # -------------------------------------------------------------------
    # Fetch data
    # -------------------------------------------------------------------
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    results = db.query(Result).filter(Result.user_id == user_id).order_by(Result.test_id).all()

    # -------------------------------------------------------------------
    # Helper: Indonesian date formatting
    # -------------------------------------------------------------------
    def id_datefmt(dt: datetime) -> str:
        days = ["Senin", "Selasa", "Rabu", "Kamis", "Jumat", "Sabtu", "Minggu"]
        months = [
            "Januari", "Februari", "Maret", "April", "Mei", "Juni",
            "Juli", "Agustus", "September", "Oktober", "November", "Desember"
        ]
        return f"{days[dt.weekday()]}, {dt.day} {months[dt.month-1]} {dt.year}"

    # -------------------------------------------------------------------
    # Compute overall test date (earliest completion)
    # -------------------------------------------------------------------
    test_dates = [r.completed_at for r in results if r.completed_at]
    overall_test_date = min(test_dates) if test_dates else datetime.now()

    # -------------------------------------------------------------------
    # Prepare meta
    # -------------------------------------------------------------------
    report_date = datetime.now()
    report_id   = f"RPT-{user_id}-{report_date.strftime('%Y%m%d%H%M')}"

    # -------------------------------------------------------------------
    # Helper functions (ratings, max scores, etc.)
    # -------------------------------------------------------------------
    def get_max_score(test_code: str) -> int | None:
        mapping = {"DISC": 24, "SPEED": 100, "MEM": 100, "LOGIC": 100, "TEMP": None, "LEAD": None}
        return mapping.get(test_code)

    def get_rating(score: int, max_score: int | None, test_name: str) -> Dict[str, str]:
        """Returns dict: {label, class, desc}"""
        if max_score is None or score is None:
            return {"label": "-", "class": "neutral", "desc": ""}
        pct = (score / max_score) * 100

        if test_name in ["Memory Test", "IQ Test", "Speed Test", "Logic & Arithmetic Test"]:
            if pct >= 80:
                return {"label": "Sangat Baik", "class": "excellent", "desc": "Performansi sangat tinggi – di atas standar."}
            if pct >= 60:
                return {"label": "Baik", "class": "good", "desc": "Performansi baik – memenuhi standar yang diharapkan."}
            if pct >= 40:
                return {"label": "Cukup", "class": "fair", "desc": "Performansi cukup – perlu pengembangan lebih lanjut."}
            return {"label": "Kurang", "class": "poor", "desc": "Performansi di bawah standar – memerlukan perhatian khusus."}
        return {"label": "-", "class": "neutral", "desc": ""}

    # -------------------------------------------------------------------
    # Sort results in the desired order
    # -------------------------------------------------------------------
    test_order = {
        "DISC": 1,
        "TEMP": 2,
        "LEAD": 3,
        "MEM": 4,
        "LOGIC": 5,
        "IQ": 6,
        "SPEED": 7
    }
    sorted_results = sorted(results, key=lambda r: test_order.get(r.test.code, 999))

    # -------------------------------------------------------------------
    # Build HTML with compact styling (removed extra meta and test-code)
    # -------------------------------------------------------------------
    html_content = f"""
<!DOCTYPE html>
<html lang="id">
<head>
<meta charset="UTF-8">
<title>Laporan Psikotes – {user.full_name or user.username}</title>
<style>
    :root {{
        --primary:   #1e3a8a;
        --secondary: #3b82f6;
        --accent:    #64748b;
        --bg:        #ffffff;
        --paper:     #f8fafc;
        --text:      #1f2937;
        --border:    #e2e8f0;
        --shadow:    0px 1px 3px rgba(15,23,42,0.08);
        --excellent: #10b981;
        --good:      #3b82f6;
        --fair:      #f59e0b;
        --poor:      #ef4444;
        --neutral:   #94a3b8;
    }}
    @page {{
        size: A4;
        margin: 2cm;  /* reduced from 2.5cm to save space */
        @top-center {{
            content: "LAPORAN PSIKOTES – {user.full_name or user.username}";
            font-size: 9pt;
            color: var(--accent);
            border-bottom: 1px solid var(--border);
            padding-bottom: 4px;
        }}
        @bottom-center {{
            content: "Halaman " counter(page) " dari " counter(pages);
            font-size: 9pt;
            color: var(--accent);
        }}
    }}
    body {{
        font-family: "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif;
        font-size: 10.5pt;  /* slightly reduced */
        line-height: 1.5;
        color: var(--text);
        background: var(--bg);
        margin: 0;
        padding: 0;
    }}
    .report-header {{
        text-align: center;
        margin-bottom: 1.5rem;
        padding-bottom: 0.75rem;
        border-bottom: 2px solid var(--primary);
    }}
    .report-header h1 {{
        font-size: 18pt;  /* slightly smaller */
        font-weight: 700;
        color: var(--primary);
        margin: 0 0 0.2rem 0;
        letter-spacing: -0.02em;
    }}
    .report-header .subtitle {{
        font-size: 9.5pt;
        color: var(--accent);
        margin: 0;
        font-weight: 500;
    }}
    .profile-card {{
        background: var(--paper);
        border: 1px solid var(--border);
        border-radius: 6px;
        padding: 1rem 1.25rem;
        margin-bottom: 1.25rem;
        box-shadow: var(--shadow);
        display: grid;
        grid-template-columns: repeat(3, 1fr);
        gap: 0.75rem;
    }}
    .profile-item {{
        display: flex;
        flex-direction: column;
    }}
    .profile-label {{
        font-size: 8pt;
        text-transform: uppercase;
        letter-spacing: 0.6px;
        color: var(--accent);
        margin-bottom: 0.2rem;
        font-weight: 600;
    }}
    .profile-value {{
        font-size: 10.5pt;
        font-weight: 700;
        color: var(--text);
    }}
    .test-card {{
        background: var(--bg);
        border: 1px solid var(--border);
        border-radius: 6px;
        padding: 1rem 1.25rem;
        margin-bottom: 1.25rem;
        box-shadow: var(--shadow);
        page-break-inside: avoid;
    }}
    .test-head {{
        border-bottom: 2px solid var(--secondary);
        padding-bottom: 0.4rem;
        margin-bottom: 0.8rem;
    }}
    .test-title {{
        font-size: 12pt;
        font-weight: 700;
        color: var(--primary);
        margin: 0;
    }}
    .narrative {{
        font-size: 10.2pt;
        text-align: justify;
        margin-bottom: 0.8rem;
        color: #334155;
    }}
    table {{
        width: 100%;
        border-collapse: collapse;
        font-size: 10pt;
        margin: 0.8rem 0;
        background: var(--bg);
    }}
    th {{
        background: var(--primary);
        color: #fff;
        font-weight: 600;
        padding: 0.4rem 0.6rem;
        text-align: left;
        border: 1px solid var(--primary);
    }}
    td {{
        padding: 0.4rem 0.6rem;
        border: 1px solid var(--border);
    }}
    tr:nth-child(even) {{
        background: var(--paper);
    }}
    tr.primary-row {{
        background: #eff6ff !important;
        font-weight: 700;
        color: var(--primary);
    }}
    tr.primary-row td:first-child {{
        border-left: 3px solid var(--secondary);
    }}
    .conclusion-cell {{
        background: #f0f9ff;
        font-style: italic;
        color: #1e40af;
        padding: 0.6rem;
        vertical-align: middle;
    }}
    .score-row {{
        display: flex;
        align-items: center;
        margin-bottom: 0.4rem;
        font-size: 10pt;
    }}
    .score-label {{
        flex: 0 0 120px;
        font-weight: 600;
        color: var(--text);
    }}
    .score-bar-wrap {{
        flex: 1;
        height: 8px;
        background: var(--border);
        border-radius: 4px;
        overflow: hidden;
        margin: 0 0.5rem;
    }}
    .score-bar {{
        height: 100%;
        background: var(--secondary);
    }}
    .score-value {{
        flex: 0 0 50px;
        text-align: right;
        font-weight: 700;
        color: var(--primary);
    }}
    .rating-badge {{
        display: inline-block;
        padding: 2px 8px;
        border-radius: 10px;
        font-size: 8.5pt;
        font-weight: 700;
        text-transform: uppercase;
        letter-spacing: 0.5px;
        margin-left: 0.3rem;
    }}
    .rating-excellent {{ background: #d1fae5; color: #065f46; }}
    .rating-good      {{ background: #dbeafe; color: #1e40af; }}
    .rating-fair      {{ background: #fef3c7; color: #92400e; }}
    .rating-poor      {{ background: #fee2e2; color: #991b1b; }}
    .rating-neutral   {{ background: #f1f5f9; color: #475569; }}
    .report-footer {{
        margin-top: 2rem;
        padding-top: 0.75rem;
        border-top: 1px solid var(--border);
        font-size: 8.5pt;
        color: var(--accent);
        text-align: center;
        font-weight: 500;
    }}
</style>
</head>
<body>

<div class="report-header">
    <h1>LAPORAN HASIL PSIKOTES</h1>
    <div class="subtitle">Psychological Assessment Report – Andamas Standard</div>
</div>

<!-- PARTICIPANT PROFILE -->
<div class="profile-card">
    <div class="profile-item">
        <span class="profile-label">Nama Lengkap</span>
        <span class="profile-value">{user.full_name or '-'}</span>
    </div>
    <div class="profile-item">
        <span class="profile-label">Pendidikan</span>
        <span class="profile-value">{user.education or '-'}</span>
    </div>
    <div class="profile-item">
        <span class="profile-label">Usia</span>
        <span class="profile-value">{user.age or '-'} Tahun</span>
    </div>
    <div class="profile-item">
        <span class="profile-label">Departemen</span>
        <span class="profile-value">{user.department or '-'}</span>
    </div>
    <div class="profile-item">
        <span class="profile-label">Posisi</span>
        <span class="profile-value">{user.position or '-'}</span>
    </div>
    <div class="profile-item">
        <span class="profile-label">Tanggal Tes</span>
        <span class="profile-value">{id_datefmt(overall_test_date)}</span>
    </div>
</div>

<!-- TEST RESULTS -->
"""

    for r in sorted_results:
        test_code = r.test.code
        test_name = r.test.name
        details   = r.details or {}

        html_content += f"""
<div class="test-card">
    <div class="test-head">
        <h2 class="test-title">{test_name}</h2>
    </div>
"""

        # ------------------------------ DISC ------------------------------
        if test_code == "DISC" and details:
            traits = ['D','I','S','C']
            graph_ii = details.get('graph_ii', {})
            graph_i = details.get('graph_i', {})
            stress_gap = details.get('stress_gap', 0)
            nat_primary = max(traits, key=lambda t: graph_ii.get(t, 0))
            pre_primary = max(traits, key=lambda t: graph_i.get(t, 0))
            trait_names = {'D':'Dominance','I':'Influence','S':'Steadiness','C':'Conscientiousness'}

            narrative = (
                f"Berdasarkan hasil tes DISC, testee menunjukkan perilaku utama alami yang dominan {trait_names[nat_primary]} ({nat_primary}). "
                f"Artinya testee cenderung stabil, sabar, dan pendengar yang baik. Mereka berfokus pada kerja sama tim, menjaga keharmonisan, serta lebih nyaman bekerja di lingkungan yang teratur dan konsisten. "
                f"Saat di bawah tekanan, testee cenderung lebih keras, tidak sabar, bahkan mendominasi. Mereka ingin segera menyelesaikan masalah dengan cepat, meskipun kadang mengabaikan detail atau perasaan orang lain (Dominance - D). "
                f"Secara alami (True Self), testee cenderung sabar, konsisten, dan setia. Mereka menghargai kestabilan, serta lebih suka bekerja dalam lingkungan yang harmonis dan minim konflik (Steadiness - S). "
                f"Berdasarkan hasil tes DISC, testee menunjukkan profil kepribadian yang Cukup Sesuai dengan Norm Standard yang telah ditetapkan oleh Andamas. Dengan demikian, testee dinilai memenuhi ketentuan yang dibutuhkan untuk level posisi saat ini dari sisi kepribadian DISC."
            )
            html_content += f'<div class="narrative">{narrative}</div>'

        # -------------------------- TEMPERAMENT --------------------------
        elif test_code == "TEMP" and details:
            raw_scores = details.get('raw_scores', {})
            primary_en = details.get('primary', '')
            
            en_to_id = {
                "Choleric": "Koleris",
                "Melancholic": "Melankolis",
                "Sanguine": "Sanguin",
                "Phlegmatic": "Plegmatis"
            }
            primary_id = en_to_id.get(primary_en, primary_en)
            
            categories = [
                ('Koleris', raw_scores.get('C', 0)),
                ('Melankolis', raw_scores.get('M', 0)),
                ('Sanguin', raw_scores.get('S', 0)),
                ('Plegmatis', raw_scores.get('P', 0)),
            ]
            
            trait_meanings = {
                'Koleris': "individu yang bersemangat, tegas, dan berorientasi pada pencapaian. Mereka pemimpin alami yang cepat mengambil keputusan namun perlu waspada terhadap kesabaran terhadap tim.",
                'Melankolis': "individu yang analitis, detail‑orientated, dan perfeksionis. Mereka menghasilkan pekerjaan berkualitas tinggi, namun perlu dukungan untuk fleksibilitas dan delegasi.",
                'Sanguin': "individu yang sosial, optimis, dan kreatif. Mereka mampu membangun rapport yang kuat, namun memerlukan struktur untuk menjaga konsistensi.",
                'Plegmatis': "individu yang stabil, sabar, dan kooperatif. Mereka menjadi penjaga keharmonisan tim, namun perlu dorongan untuk inisiatif dan kecepatan kerja."
            }
            # Build conclusion text
            concl_text = f"Berdasarkan hasil tes, testee menunjukkan dominasi temperament {primary_id}, yaitu {trait_meanings.get(primary_id, '')}"

            # Start table with conclusion cell spanning all rows
            html_content += '<table>'
            html_content += '<tr><th>Kategori</th><th>Skor</th><th>Kesimpulan</th></tr>'
            rowspan = len(categories)
            for i, (cat, score) in enumerate(categories):
                is_primary = (primary_id == cat)
                row_class = ' class="primary-row"' if is_primary else ''
                if i == 0:
                    # First row gets the conclusion with rowspan
                    html_content += f'<tr{row_class}><td>{cat}</td><td>{score}</td><td class="conclusion-cell" rowspan="{rowspan}">{concl_text}</td></tr>'
                else:
                    html_content += f'<tr{row_class}><td>{cat}</td><td>{score}</td></tr>'
            html_content += '</table>'

        # -------------------------- LEADERSHIP --------------------------
        elif test_code == "LEAD" and details:
            score = r.score or 0
            html_content += f"""
            <div class="narrative">
                Hasil Psikotes menunjukkan bahwa testee menjawab soal dengan skor keseluruhan <strong>{score} Point</strong>. Dengan demikian dapat dikatakan bahwa testee “menunjukkan motivasi yang kuat untuk menjadi pemimpin”.
            </div>
            """

        # -------------------------- MEMORY -----------------------------
        elif test_code == "MEM" and details:
            correct = details.get('correct_count', 0)
            total = 50
            rating = get_rating(correct, total, test_name)
            pct = int((correct / total) * 100)
            html_content += f"""
            <div class="score-row">
                <span class="score-label">Kemampuan Memori</span>
                <div class="score-bar-wrap">
                    <div class="score-bar" style="width:{pct}%; background:var(--{rating['class']});"></div>
                </div>
                <span class="score-value">{correct}/{total}</span>
            </div>
            <div class="narrative">
                Testee menjawab <strong>{correct}</strong> soal benar dari total {total} soal. 
                Dengan demikian, kemampuan memori dinilai 
                <span class="rating-badge rating-{rating['class']}">{rating['label']}</span>.
                <br>{rating['desc']}
            </div>
            """

        # ----------------------- LOGIC & ARITHMETIC --------------------
        elif test_code == "LOGIC" and details:
            correct = details.get('correct_count', 0)
            total = 25
            rating = get_rating(correct, total, test_name)
            pct = int((correct / total) * 100)
            html_content += f"""
            <div class="score-row">
                <span class="score-label">Logika & Aritmatika</span>
                <div class="score-bar-wrap">
                    <div class="score-bar" style="width:{pct}%; background:var(--{rating['class']});"></div>
                </div>
                <span class="score-value">{correct}/{total}</span>
            </div>
            <div class="narrative">
                Testee menjawab <strong>{correct}</strong> soal benar dari {total} soal. 
                Kemampuan logika dan pemecahan masalah dinilai 
                <span class="rating-badge rating-{rating['class']}">{rating['label']}</span>.
                <br>{rating['desc']}
            </div>
            """

        # --------------------------- SPEED -----------------------------
        elif test_code == "SPEED" and details:
            correct = details.get('score', 0)   # Use 'score' field
            total = 100
            rating = get_rating(correct, total, test_name)
            pct = int((correct / total) * 100)
            extra = "Unggul dalam Tekanan Tinggi" if rating['class'] == 'excellent' else ""
            html_content += f"""
            <div class="score-row">
                <span class="score-label">Kecepatan & Akurasi</span>
                <div class="score-bar-wrap">
                    <div class="score-bar" style="width:{pct}%; background:var(--{rating['class']});"></div>
                </div>
                <span class="score-value">{correct}/{total}</span>
            </div>
            <div class="narrative">
                Dari {total} soal Speed Test, testee menjawab <strong>{correct}</strong> soal benar. 
                {('Dengan performa ini, testee dinilai “<strong>'+extra+'</strong>” – unggul dan berkembang di bawah tekanan tinggi. Ideal untuk peran dengan tenggat waktu ketat dan situasi kritis.' if extra else 'Dengan demikian kemampuan kecepatan dan akurasi dinilai <span class="rating-badge rating-'+rating['class']+'">'+rating['label']+'</span>.')}
            </div>
            """

        # --------------------------- IQ TEST (placeholder) -------------
        elif test_code == "IQ" and details:
            correct = details.get('correct', 0)
            total = 20
            rating = get_rating(correct, total, test_name)
            pct = int((correct / total) * 100)
            html_content += f"""
            <div class="score-row">
                <span class="score-label">Kemampuan Pola (IQ)</span>
                <div class="score-bar-wrap">
                    <div class="score-bar" style="width:{pct}%; background:var(--{rating['class']});"></div>
                </div>
                <span class="score-value">{correct}/{total}</span>
            </div>
            <div class="narrative">
                Testee berhasil menjawab <strong>{correct}</strong> soal benar dari {total} soal. 
                Kemampuan memahami pola dan berpikir abstrak dinilai 
                <span class="rating-badge rating-{rating['class']}">{rating['label']}</span>.
            </div>
            """

        # ------------------------ FALLBACK ---------------------------
        else:
            max_sc = get_max_score(test_code)
            html_content += f'<div class="narrative">Skor: <strong>{r.score or 0}</strong> / {max_sc if max_sc else "N/A"}</div>'

        html_content += "</div>"  # close test-card

    html_content += f"""
<div class="report-footer">
    <strong>Laporan Dikembangkan oleh Tim Psikologi Andamas</strong> | 
    Data bersifat rahasia dan hanya untuk penggunaan internal. 
    <br>Dicetak pada {id_datefmt(report_date)} – {report_id}
</div>
</body>
</html>
"""

    # -------------------------------------------------------------------
    # Generate PDF
    # -------------------------------------------------------------------
    pdf_bytes = HTML(string=html_content).write_pdf()

    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="Laporan_Psikotes_{user.username or user.id}_{report_date.strftime("%Y%m%d")}.pdf"'
        },
    )