"""
Admin routes - Statistics, logs, and bulk operations
"""
from fastapi import APIRouter, Depends, HTTPException, Form, UploadFile, File
from sqlalchemy.orm import Session
from sqlalchemy import func, desc
from typing import Optional
from datetime import datetime, date
import csv
import io
import secrets
import string

from auth import require_admin, require_superadmin, hash_password
from database import get_db
from models import User, Test, Assignment, Result, ExitLog, Question, ClassConfig
from utils import get_max_score

router = APIRouter(tags=["admin"])


# ==================== STATISTICS ====================

@router.get("/admin/stats/summary")
def get_stats_summary(db: Session = Depends(get_db), admin: User = Depends(require_admin)):
    """Get overall statistics summary (admin and superadmin)"""
    total_participants = db.query(User).filter(User.role == "participant").count()
    total_assignments = db.query(Assignment).count()
    completed_assignments = db.query(Assignment).filter(Assignment.status == "completed").count()
    completion_rate = (completed_assignments / total_assignments * 100) if total_assignments else 0
    avg_score = db.query(func.avg(Result.score)).scalar() or 0

    return {
        "total_participants": total_participants,
        "total_assignments": total_assignments,
        "completed_assignments": completed_assignments,
        "completion_rate": round(completion_rate, 2),
        "average_score": round(avg_score, 2)
    }


@router.get("/admin/stats/tests")
def get_test_stats(db: Session = Depends(get_db), admin: User = Depends(require_admin)):
    """Get statistics per test (admin and superadmin)"""
    tests = db.query(Test).all()
    result = []
    for test in tests:
        completed_count = db.query(Assignment).filter(
            Assignment.test_id == test.id,
            Assignment.status == "completed"
        ).count()
        avg_score = db.query(func.avg(Result.score)).filter(Result.test_id == test.id).scalar() or 0
        result.append({
            "test_id": test.id,
            "test_name": test.name,
            "completed_count": completed_count,
            "avg_score": round(avg_score, 2),
            "max_score": get_max_score(test.code)
        })
    return result


@router.get("/admin/stats/recent")
def get_recent_activity(
    limit: int = 10,
    db: Session = Depends(get_db),
    admin: User = Depends(require_admin)
):
    """Get recent test completions (admin and superadmin)"""
    recent_results = db.query(Result).order_by(desc(Result.completed_at)).limit(limit).all()
    output = []
    for r in recent_results:
        total_questions = db.query(Question).filter(Question.test_id == r.test_id).count()
        output.append({
            "id": r.id,
            "participant_name": r.user.full_name or r.user.username,
            "test_name": r.test.name,
            "score": r.score,
            "total_questions": total_questions,
            "completed_at": r.completed_at.isoformat() if r.completed_at else None
        })
    return output


@router.get("/admin/stats/completion")
def get_completion_stats(db: Session = Depends(get_db), admin: User = Depends(require_admin)):
    """Get assignment completion statistics (admin and superadmin)"""
    total = db.query(Assignment).count()
    completed = db.query(Assignment).filter(Assignment.status == "completed").count()
    in_progress = db.query(Assignment).filter(Assignment.status == "in_progress").count()
    pending = db.query(Assignment).filter(Assignment.status == "pending").count()
    locked = db.query(Assignment).filter(Assignment.status == "locked").count()
    
    # Get incomplete submissions (answered < total questions)
    incomplete_count = 0
    results = db.query(Result).all()
    for r in results:
        if r.details and isinstance(r.details, dict):
            if r.details.get("is_complete") == False:
                incomplete_count += 1
    
    return {
        "total": total,
        "completed": completed,
        "in_progress": in_progress,
        "pending": pending,
        "locked": locked,
        "incomplete_submissions": incomplete_count
    }


@router.get("/admin/stats/security-events")
def get_security_events(
    limit: int = 10,
    db: Session = Depends(get_db),
    admin: User = Depends(require_admin)
):
    """Get recent security events (locked assignments, exit logs) (admin and superadmin)"""
    # Get recent exit logs
    exit_logs = db.query(ExitLog).order_by(desc(ExitLog.timestamp)).limit(limit).all()
    
    # Get locked assignments with user info
    locked_assignments = db.query(Assignment).filter(
        Assignment.status == "locked"
    ).order_by(desc(Assignment.assigned_at)).limit(limit).all()
    
    events = []
    
    # Add exit log events
    for log in exit_logs:
        events.append({
            "type": "exit_log",
            "timestamp": log.timestamp.isoformat() if log.timestamp else None,
            "username": log.user.username if log.user else "Unknown",
            "full_name": log.user.full_name if log.user else "Unknown",
            "test_name": log.assignment.test.name if log.assignment and log.assignment.test else "Unknown",
            "reason": "Exit attempt detected",
            "severity": "warning"
        })
    
    # Add locked assignment events
    for a in locked_assignments:
        events.append({
            "type": "locked",
            "timestamp": a.assigned_at.isoformat() if a.assigned_at else None,
            "username": a.user.username,
            "full_name": a.user.full_name,
            "test_name": a.test.name,
            "reason": "Test locked (security violation)",
            "severity": "critical"
        })
    
    # Sort by timestamp descending
    events.sort(key=lambda x: x["timestamp"] or "", reverse=True)
    
    return events[:limit]


# ==================== SECURITY & LOGS ====================

@router.get("/admin/locked-assignments")
def get_locked_assignments(db: Session = Depends(get_db), superadmin: User = Depends(require_superadmin)):
    """Get all locked assignments (superadmin only)"""
    assignments = db.query(Assignment).filter(Assignment.status == "locked").all()
    result = []
    for a in assignments:
        result.append({
            "id": a.id,
            "user_id": a.user_id,
            "username": a.user.username,
            "full_name": a.user.full_name,
            "test_name": a.test.name,
            "status": a.status,
            "assigned_at": a.assigned_at,
        })
    return result


@router.get("/admin/exit-logs")
def get_exit_logs(
    user_id: Optional[int] = None,
    assignment_id: Optional[int] = None,
    from_date: Optional[date] = None,
    to_date: Optional[date] = None,
    db: Session = Depends(get_db),
    superadmin: User = Depends(require_superadmin)
):
    """Get exit logs with filters (superadmin only)"""
    query = db.query(ExitLog).join(User).join(Assignment)
    if user_id:
        query = query.filter(ExitLog.user_id == user_id)
    if assignment_id:
        query = query.filter(ExitLog.assignment_id == assignment_id)
    if from_date:
        query = query.filter(ExitLog.timestamp >= from_date)
    if to_date:
        query = query.filter(ExitLog.timestamp <= to_date)
    logs = query.order_by(ExitLog.timestamp.desc()).all()
    output = []
    for log in logs:
        output.append({
            "id": log.id,
            "user_id": log.user_id,
            "username": log.user.username,
            "full_name": log.user.full_name,
            "assignment_id": log.assignment_id,
            "test_name": log.assignment.test.name,
            "timestamp": log.timestamp
        })
    return output


# ==================== BULK OPERATIONS ====================

@router.post("/admin/users/bulk")
def bulk_create_users(
    file: UploadFile = File(...),
    assign_all: bool = Form(False),
    db: Session = Depends(get_db),
    admin: User = Depends(require_admin)
):
    """Bulk create users from CSV or Excel file"""
    filename = file.filename.lower()
    if not (filename.endswith('.csv') or filename.endswith('.xlsx') or filename.endswith('.xls')):
        raise HTTPException(status_code=400, detail="Only CSV or Excel files are allowed")

    contents = file.file.read()
    rows = []

    if filename.endswith('.csv'):
        decoded = contents.decode('utf-8-sig')
        csv_reader = csv.DictReader(io.StringIO(decoded))
        rows = list(csv_reader)

    elif filename.endswith('.xlsx'):
        from openpyxl import load_workbook
        from io import BytesIO
        wb = load_workbook(filename=BytesIO(contents), read_only=True)
        sheet = wb.active
        all_rows = list(sheet.iter_rows(values_only=True))
        if not all_rows:
            raise HTTPException(status_code=400, detail="Excel file is empty")
        headers = [str(cell) if cell is not None else '' for cell in all_rows[0]]
        rows = []
        for row in all_rows[1:]:
            row_dict = {}
            for idx, val in enumerate(row):
                if idx < len(headers):
                    row_dict[headers[idx]] = str(val) if val is not None else ''
            rows.append(row_dict)

    elif filename.endswith('.xls'):
        import xlrd
        from io import BytesIO
        book = xlrd.open_workbook(file_contents=contents)
        sheet = book.sheet_by_index(0)
        headers = [str(sheet.cell_value(0, col)) for col in range(sheet.ncols)]
        rows = []
        for row_idx in range(1, sheet.nrows):
            row_dict = {}
            for col_idx in range(sheet.ncols):
                val = sheet.cell_value(row_idx, col_idx)
                row_dict[headers[col_idx]] = str(val) if val else ''
            rows.append(row_dict)

    required_fields = ['username', 'password', 'full_name', 'class', 'level']
    VALID_LEVELS = [
        "Operator / Mekanik", 
        "Admin / Non - Staff", 
        "Foreman / Officer", 
        "Supervisor / Section Head", 
        "Superintendent / Dept. Head"
    ]
    results = {
        "total": len(rows),
        "success": 0,
        "failed": 0,
        "errors": []
    }
    created_users = []

    for row_num, row in enumerate(rows, start=2):
        # Skip empty rows (common in Excel/CSV exports)
        if not any(str(v).strip() for v in row.values()) or not str(row.get('username', '')).strip():
            continue

        # Allow 'kelas' as an alias for 'class' in the CSV

        if 'kelas' in row and not row.get('class'):
            row['class'] = row['kelas']
            
        # Allow 'unit_bisnis' or 'unit' as aliases for 'business_unit'
        business_unit_val = row.get('unit_bisnis') or row.get('unit') or row.get('business_unit')
            
        missing = [f for f in required_fields if not str(row.get(f, '')).strip()]
        if missing:
            results["failed"] += 1
            results["errors"].append(f"Baris {row_num}: Kolom wajib tidak ada/kosong: {', '.join(missing)}")
            continue

        username = str(row['username']).strip()
        password = str(row['password']).strip()
        full_name = str(row['full_name']).strip()
        class_name = str(row['class']).strip()
        level_input = str(row['level']).strip()

        # Normalize level (case-insensitive)
        level = next((l for l in VALID_LEVELS if l.lower() == level_input.lower()), None)
        if not level:
            results["failed"] += 1
            results["errors"].append(f"Baris {row_num}: Level '{level_input}' tidak valid. Gunakan salah satu dari: {', '.join(VALID_LEVELS)}")
            continue

        if db.query(User).filter(User.username == username).first():
            results["failed"] += 1
            results["errors"].append(f"Baris {row_num}: Username '{username}' sudah terdaftar")
            continue

        # Resolve class_id from class name
        class_config = db.query(ClassConfig).filter(
            ClassConfig.name.ilike(class_name)
        ).first()

        if not class_config:
            results["failed"] += 1
            results["errors"].append(f"Baris {row_num}: Klasifikasi '{class_name}' tidak ditemukan")
            continue

        user_data = {
            'username': username,
            'full_name': full_name,
            'age': row.get('age') if str(row.get('age', '')).strip() else None,
            'gender': str(row.get('gender', '')).strip() or None,
            'education': str(row.get('education', '')).strip() or None,
            'department': str(row.get('department', '')).strip() or None,
            'position': str(row.get('position', '')).strip() or None,
            'business_unit': str(business_unit_val).strip() if business_unit_val else None,
            'level': level,
            'role': 'participant'
        }

        if user_data['age']:
            try:
                user_data['age'] = int(float(str(user_data['age']))) # Handle float-like strings from Excel
            except (ValueError, TypeError):
                results["failed"] += 1
                results["errors"].append(f"Baris {row_num}: Usia harus berupa angka (menemukan: '{user_data['age']}')")
                continue

        new_user = User(
            username=user_data['username'],
            password_hash=hash_password(password),
            role=user_data['role'],
            full_name=user_data['full_name'],
            age=user_data['age'],
            gender=user_data['gender'],
            education=user_data['education'],
            department=user_data['department'],
            position=user_data['position'],
            business_unit=user_data['business_unit'],
            level=user_data['level'],
            class_id=class_config.id
        )
        db.add(new_user)
        db.flush()
        created_users.append(new_user)
        results["success"] += 1

    if assign_all and created_users:
        all_tests = db.query(Test).all()
        for user in created_users:
            for test in all_tests:
                existing = db.query(Assignment).filter(
                    Assignment.user_id == user.id,
                    Assignment.test_id == test.id
                ).first()
                if not existing:
                    new_assignment = Assignment(user_id=user.id, test_id=test.id)
                    db.add(new_assignment)

    db.commit()
    return results
